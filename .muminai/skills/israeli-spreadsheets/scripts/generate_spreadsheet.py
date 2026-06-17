#!/usr/bin/env python3
"""Generate Israeli financial Excel spreadsheets.

Creates formatted Hebrew spreadsheets with Israeli tax calculations,
VAT, NIS formatting, and common financial templates.

Usage:
    python generate_spreadsheet.py --template invoice --output invoice.xlsx
    python generate_spreadsheet.py --template salary --output salary.xlsx
    python generate_spreadsheet.py --template arnona --output arnona.xlsx
    python generate_spreadsheet.py --help

Requirements:
    pip install openpyxl
"""

import argparse
import sys
from decimal import Decimal

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Israeli financial constants (2026, VAT at 18% since 2025-01-01)
# Income tax brackets reflect Amendment 288 (effective 1 January 2026):
# the 20% and 31% bands were widened, pushing the 35% floor from 269,280
# to 301,200 NIS annually. The 10/14/35/47/50 rate thresholds are otherwise
# frozen at 2025 levels through 2027.
VAT_RATE = Decimal("0.18")
TAX_BRACKETS = [
    (Decimal("84120"), Decimal("0.10")),
    (Decimal("120720"), Decimal("0.14")),
    (Decimal("228000"), Decimal("0.20")),
    (Decimal("301200"), Decimal("0.31")),
    (Decimal("560280"), Decimal("0.35")),
    (Decimal("721560"), Decimal("0.47")),
    (Decimal("999999999"), Decimal("0.50")),
]
CREDIT_POINT_VALUE = Decimal("2904")
RESIDENT_CREDIT_POINTS = Decimal("2.25")

NIS_FORMAT = '#,##0.00 "₪"'
PERCENT_FORMAT = "0.00%"


def setup_rtl_sheet(ws, title):
    """Configure worksheet for RTL Hebrew display."""
    ws.title = title
    ws.sheet_view.rightToLeft = True


def style_header(ws, row, cols, fill_color="1F4E79"):
    """Style header row with colors."""
    header_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    header_font = Font(name="Heebo", size=11, bold=True, color="FFFFFF")
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def create_invoice(output_path):
    """Create a Hebrew tax invoice template."""
    wb = Workbook()
    ws = wb.active
    setup_rtl_sheet(ws, "חשבונית מס")

    # Business details
    ws["A1"] = "חשבונית מס / קבלה"
    ws["A1"].font = Font(name="Heebo", size=16, bold=True)
    ws["A2"] = "שם העסק: [שם]"
    ws["A3"] = "ע.מ./ח.פ.: [מספר]"
    ws["A4"] = "כתובת: [כתובת]"
    ws["A5"] = "טלפון: [טלפון]"

    # Invoice details
    ws["D2"] = "מספר חשבונית:"
    ws["E2"] = "[מספר]"
    ws["D3"] = "תאריך:"
    ws["E3"] = "[DD/MM/YYYY]"

    # Item headers
    headers = ["תיאור", "כמות", "מחיר ליחידה", "סה\"כ"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=7, column=i, value=h)
    style_header(ws, 7, 4)

    # Sample rows
    for row in range(8, 11):
        ws.cell(row=row, column=1, value="[פריט]")
        ws.cell(row=row, column=2, value=1)
        ws.cell(row=row, column=3, value=0).number_format = NIS_FORMAT
        ws.cell(row=row, column=4).number_format = NIS_FORMAT

    # Totals
    ws.cell(row=12, column=3, value="סכום ביניים:").font = Font(bold=True)
    ws.cell(row=12, column=4).number_format = NIS_FORMAT
    ws.cell(row=13, column=3, value="מע\"מ (18%):").font = Font(bold=True)
    ws.cell(row=13, column=4).number_format = NIS_FORMAT
    ws.cell(row=14, column=3, value="סה\"כ לתשלום:").font = Font(name="Heebo", bold=True, size=12)
    ws.cell(row=14, column=4).number_format = NIS_FORMAT

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    wb.save(output_path)
    print(f"Created invoice template: {output_path}")


def calculate_income_tax(annual_income):
    """Calculate Israeli progressive income tax."""
    tax = Decimal("0")
    prev_limit = Decimal("0")
    for limit, rate in TAX_BRACKETS:
        if annual_income <= prev_limit:
            break
        taxable = min(annual_income, limit) - prev_limit
        tax += taxable * rate
        prev_limit = limit

    credit = RESIDENT_CREDIT_POINTS * CREDIT_POINT_VALUE
    return max(Decimal("0"), tax - credit)


def create_salary_slip(output_path):
    """Create a Hebrew salary slip (tlush maskoret) template."""
    wb = Workbook()
    ws = wb.active
    setup_rtl_sheet(ws, "תלוש משכורת")

    ws["A1"] = "תלוש משכורת"
    ws["A1"].font = Font(name="Heebo", size=16, bold=True)

    ws["A3"] = "שם העובד:"
    ws["B3"] = "[שם]"
    ws["A4"] = "ת.ז.:"
    ws["B4"] = "[מספר]"
    ws["A5"] = "חודש:"
    ws["B5"] = "[חודש/שנה]"

    # Earnings
    ws["A7"] = "תשלומים"
    ws["A7"].font = Font(bold=True)
    style_header(ws, 7, 2, "2E7D32")

    earnings = [
        ("שכר בסיס", 0),
        ("שעות נוספות", 0),
        ("בונוס", 0),
    ]
    for i, (label, amount) in enumerate(earnings, 8):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=amount).number_format = NIS_FORMAT

    # Deductions
    ws["A12"] = "ניכויים"
    ws["A12"].font = Font(bold=True)
    style_header(ws, 12, 2, "C62828")

    deductions = [
        ("מס הכנסה", 0),
        ("ביטוח לאומי", 0),
        ("מס בריאות", 0),
        ("פנסיה עובד", 0),
        ("קרן השתלמות עובד", 0),
    ]
    for i, (label, amount) in enumerate(deductions, 13):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=amount).number_format = NIS_FORMAT

    # Net
    ws.cell(row=19, column=1, value="שכר נטו").font = Font(name="Heebo", bold=True, size=14)
    ws.cell(row=19, column=2).number_format = NIS_FORMAT

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 18

    wb.save(output_path)
    print(f"Created salary slip: {output_path}")


def create_arnona(output_path):
    """Create an arnona calculator spreadsheet."""
    wb = Workbook()
    ws = wb.active
    setup_rtl_sheet(ws, "מחשבון ארנונה")

    ws["A1"] = "מחשבון ארנונה"
    ws["A1"].font = Font(name="Heebo", size=16, bold=True)

    headers = ["עיר", "תעריף למ\"ר (דו-חודשי)", "שטח (מ\"ר)", "ארנונה דו-חודשית", "ארנונה שנתית"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 5)

    cities = [
        ("תל אביב-יפו", 55.80),
        ("ירושלים", 40.50),
        ("חיפה", 33.20),
        ("באר שבע", 27.90),
        ("נתניה", 43.10),
        ("רעננה", 48.50),
        ("הרצליה", 52.30),
        ("פתח תקווה", 38.70),
        ("ראשון לציון", 41.20),
    ]

    for i, (city, rate) in enumerate(cities, 4):
        ws.cell(row=i, column=1, value=city)
        ws.cell(row=i, column=2, value=rate).number_format = NIS_FORMAT
        ws.cell(row=i, column=3, value=80)  # default 80 sqm
        ws.cell(row=i, column=4).number_format = NIS_FORMAT
        ws.cell(row=i, column=4, value=f"=B{i}*C{i}")
        ws.cell(row=i, column=5).number_format = NIS_FORMAT
        ws.cell(row=i, column=5, value=f"=D{i}*6")

    for col in ["A", "B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 22

    wb.save(output_path)
    print(f"Created arnona calculator: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate Israeli financial Excel spreadsheets"
    )
    parser.add_argument(
        "--template", choices=["invoice", "salary", "arnona"],
        default="invoice",
        help="Spreadsheet template (default: invoice)"
    )
    parser.add_argument(
        "--output", default="spreadsheet.xlsx",
        help="Output file path (default: spreadsheet.xlsx)"
    )
    args = parser.parse_args()

    generators = {
        "invoice": create_invoice,
        "salary": create_salary_slip,
        "arnona": create_arnona,
    }
    generators[args.template](args.output)


if __name__ == "__main__":
    main()
